#!/usr/bin/env python3


import numpy as np
import pandas as pd
import nbformat
import re
from nbconvert.preprocessors import ExecutePreprocessor, CellExecutionError
import openai
from openai import RateLimitError, OpenAIError
from openpyxl import Workbook
from openpyxl.styles import Font
import os
from dotenv import load_dotenv
from pprint import pprint
from openai import OpenAI
from fpdf import FPDF
from tqdm import tqdm
import sys
from joblib import Parallel, delayed
import logging
import logging
import os
from joblib import Parallel, delayed
from tqdm import tqdm
from io import StringIO
import argparse
import matplotlib
matplotlib.use('Agg')  # Use a non-interactive backend
import matplotlib.pyplot as plt
import io


# Cargar las variables de entorno desde el archivo .env
load_dotenv()

# Obtener la clave de API
openai_api_key = os.getenv("OPENAI_API_KEY")

if openai_api_key is None:
    raise ValueError("API key is not set")

# Inicializar la API de OpenAI
openai.api_key = openai_api_key

################# LISTAR NOTEBOOKS #################

def listar_notebooks(directory):
    """
    Devuelve una lista con los nombres de todos los archivos con extensión '.ipynb' en un directorio dado.
    
    Parameters:
        directory (str): La ruta al directorio donde buscar los archivos.

    Returns:
        tuple: Dos listas, la primera con los nombres de los alumnos (nombres de archivos sin extensión),
               y la segunda con los nombres de los archivos completos que cumplen los criterios.

    Raises:
        RuntimeError: Si ocurre un error crítico al intentar listar los archivos.
    """
    # Crear un logger específico para la función
    logger = logging.getLogger('listar_notebooks')
    if logger.hasHandlers():
        logger.handlers.clear()

    log_stream = logging.StreamHandler()
    logger.addHandler(log_stream)
    logger.setLevel(logging.DEBUG)

    archivos = []
    try:
        for filename in os.listdir(directory):
            if filename.endswith('.ipynb'):
                archivos.append(filename)
        
        alumnos = [archivo.replace('.ipynb', '') for archivo in archivos]
        
        if not archivos:
            warning_msg = f"No se encontraron archivos .ipynb en el directorio '{directory}'."
            logger.warning(warning_msg)
            print(warning_msg)
        
        logger.info(f"Se encontraron {len(archivos)} archivos .ipynb en el directorio {directory}.")
        return alumnos, archivos

    except Exception as e:
        error_msg = f"Error al listar archivos en el directorio {directory}: {e}"
        logger.error(error_msg)
        raise RuntimeError(error_msg)


############ CARGAR CRITERIOS ################################################################################################################

def cargar_criterios(criterios_file):
    """
    Carga los criterios desde un archivo de texto y los almacena en un diccionario.

    Parameters:
        criterios_file (str): Ruta del archivo de texto que contiene los criterios.

    Returns:
        dict: Un diccionario donde las claves son los nombres de los criterios y los valores
              son diccionarios con 'descripcion' y 'ejemplo'.

    Raises:
        RuntimeError: Si ocurre un error crítico durante la carga de criterios.
    """
    criterios = {}

    # Crear un logger específico para la función
    logger = logging.getLogger('cargar_criterios')
    if logger.hasHandlers():
        logger.handlers.clear()

    log_stream = logging.StreamHandler()
    logger.addHandler(log_stream)
    logger.setLevel(logging.DEBUG)
    
    try:
        # Intentar abrir el archivo de criterios y leer su contenido
        with open(criterios_file, 'r', encoding='utf-8') as file:
            contenido = file.read()
    except FileNotFoundError:
        error_msg = f"El archivo {criterios_file} no se encontró."
        logger.error(error_msg)
        raise RuntimeError(error_msg)
    except IOError as e:
        error_msg = f"Error al leer el archivo {criterios_file}: {e}"
        logger.error(error_msg)
        raise RuntimeError(error_msg)

    try:
        # Dividir el contenido en secciones usando '@@' como delimitador
        secciones = contenido.split('@@')

        # Validar que el contenido esté correctamente estructurado
        if len(secciones) < 3 or len(secciones) % 2 == 0:
            error_msg = f"Formato incorrecto en el archivo {criterios_file}. Verifique que cada criterio tenga nombre y detalles asociados."
            logger.error(error_msg)
            raise ValueError(error_msg)

        # Iterar sobre las secciones para extraer los criterios y sus detalles
        for i in range(1, len(secciones), 2):
            nombre_criterio = secciones[i].strip()  # Extraer y limpiar el nombre del criterio
            detalles = secciones[i + 1].strip()  # Extraer y limpiar los detalles del criterio
            partes = detalles.split("Ejemplo:")  # Dividir detalles en descripción y ejemplo
            descripcion = partes[0].replace("Descripción:", "").strip()  # Limpiar la descripción
            ejemplo = partes[1].strip() if len(partes) > 1 else ""  # Limpiar el ejemplo si existe

            # Almacenar el criterio en el diccionario
            criterios[nombre_criterio] = {"descripcion": descripcion, "ejemplo": ejemplo}

        # Si no se encontraron criterios válidos, lanzar una excepción
        if not criterios:
            error_msg = f"No se encontraron criterios válidos en el archivo {criterios_file}."
            logger.error(error_msg)
            raise ValueError(error_msg)

    except IndexError:
        error_msg = "Formato incorrecto en el archivo de criterios. Verifique la estructura del archivo."
        logger.error(error_msg)
        raise RuntimeError(error_msg)
    except ValueError as e:
        logger.error(f"Error en cargar_criterios: {e}")
        raise RuntimeError(f"Error en cargar_criterios: {e}")
    except Exception as e:
        error_msg = f"Error inesperado al procesar el archivo {criterios_file}: {e}"
        logger.error(error_msg)
        raise RuntimeError(error_msg)

    logger.info("Criterios cargados correctamente. No se encontraron errores.")
    
    return criterios


############# VERIFICAR ESTRUCTURA EXAMEN ################################################################################################################

def verifica_estructura_examen(examen_file):
    """
    Verifica que un notebook Jupyter sigue la estructura esperada para un examen.

    Parameters:
        examen_file (str): Ruta del archivo del notebook de examen.

    Returns:
        None

    Raises:
        RuntimeError: Si se detecta un error crítico en la estructura.
    """
    errores = []
    contexto_detectado = False
    ejercicio_num = 0
    se_espera_solucion = False
    codigo_encontrado = False

    # Crear un logger específico para la función
    logger = logging.getLogger(f'verifica_estructura_examen')
    if logger.hasHandlers():
        logger.handlers.clear()

    log_stream = logging.StreamHandler()
    logger.addHandler(log_stream)
    logger.setLevel(logging.DEBUG)

    try:
        # Leer el notebook
        with open(examen_file, 'r', encoding='utf-8') as f:
            notebook = nbformat.read(f, as_version=4)
        logger.info(f"Archivo del examen {examen_file} leído correctamente.")

    except FileNotFoundError:
        error_msg = f"Error: El archivo {examen_file} no se encontró."
        logger.error(error_msg)
        raise RuntimeError(error_msg)
    except Exception as e:
        error_msg = f"Error al leer el archivo {examen_file}: {e}"
        logger.error(error_msg)
        raise RuntimeError(error_msg)

    try:
        # Procesar las celdas del notebook
        for cell in notebook.cells:
            if cell.cell_type == 'markdown':
                # Si se esperaba una solución y no se encontró código, generar un error
                if se_espera_solucion and not codigo_encontrado:
                    error_msg = f"La Solución del Ejercicio {ejercicio_num} no contiene código."
                    errores.append(error_msg)
                    logger.error(error_msg)
                se_espera_solucion = False
                codigo_encontrado = False

                cell_content = cell['source'].strip()

                # Verificar Contexto
                if cell_content.startswith("## Contexto"):
                    if contexto_detectado:
                        error_msg = "Más de un '## Contexto' detectado."
                        errores.append(error_msg)
                        logger.error(error_msg)
                    contexto_detectado = True

                # Verificar Ejercicios y Criterios
                if cell_content.startswith("## Ejercicio"):
                    ejercicio_num += 1
                    se_espera_solucion = True

                    if "Criterios:" not in cell_content:
                        error_msg = f"El Ejercicio {ejercicio_num} no contiene la sección 'Criterios'."
                        errores.append(error_msg)
                        logger.error(error_msg)
                    else:
                        criterios = cell_content.split("Criterios:")[-1].strip()
                        criterios_list = criterios.split(",")

                        # Verificar si cada criterio está delimitado correctamente y no está vacío
                        for criterio in criterios_list:
                            criterio = criterio.strip()
                            if not (criterio.startswith("@@") and criterio.endswith("@@")) or len(criterio) <= 4:
                                error_msg = f"Criterios mal formateados o vacíos en el Ejercicio {ejercicio_num}"
                                errores.append(error_msg)
                                logger.error(error_msg)
                                break

            elif cell.cell_type == 'code':
                if se_espera_solucion:
                    cell_content = cell['source'].strip()

                    # Eliminar saltos de línea y espacios adicionales para una comparación más robusta
                    normalized_content = " ".join(cell_content.split())

                    # Verificar si la celda comienza con "## Solución ejercicio" o "## Solucion ejercicio"
                    if normalized_content.startswith("## Solución ejercicio") or normalized_content.startswith("## Solucion ejercicio"):
                        try:
                            sol_num = int(normalized_content.split("## Solución ejercicio" if "## Solución ejercicio" in normalized_content else "## Solucion ejercicio")[1].strip().split()[0])
                            if sol_num != ejercicio_num:
                                error_msg = f"La Solución ejercicio {sol_num} no corresponde al Ejercicio {ejercicio_num}."
                                errores.append(error_msg)
                                logger.error(error_msg)
                            codigo_encontrado = False  # Reiniciar el indicador de código encontrado
                        except (ValueError, IndexError):
                            error_msg = f"Formato de número incorrecto en la Solución del Ejercicio {ejercicio_num}."
                            errores.append(error_msg)
                            logger.error(error_msg)
                    else:
                        error_msg = f"El Ejercicio {ejercicio_num} tiene una celda de código que no comienza con '## Solución ejercicio {ejercicio_num}' o '## Solucion ejercicio {ejercicio_num}'."
                        errores.append(error_msg)
                        logger.error(error_msg)
                        se_espera_solucion = False

                    # Verificar si la celda contiene código más allá de comentarios o está vacía
                    lines = cell_content.split("\n")
                    for line in lines:
                        stripped_line = line.strip()
                        if stripped_line and not stripped_line.startswith("#"):
                            codigo_encontrado = True
                            break

        # Verificar si la última solución esperada fue proporcionada y si tenía código
        if se_espera_solucion and not codigo_encontrado:
            error_msg = f"La Solución del Ejercicio {ejercicio_num} no contiene código."
            errores.append(error_msg)
            logger.error(error_msg)

        # Verificaciones finales
        if not contexto_detectado:
            error_msg = "No se detectó un '## Contexto' en el notebook."
            errores.append(error_msg)
            logger.error(error_msg)

    except Exception as e:
        error_msg = f"Error inesperado al procesar el archivo {examen_file}: {e}"
        logger.error(error_msg)
        raise RuntimeError(error_msg)

    # Log de resultado final
    if errores:
        logger.error("EXAMEN Estructura Incorrecta: Se encontraron los siguientes errores:")
        for error in errores:
            logger.error(error)
        raise RuntimeError("EXAMEN Se encontraron errores críticos en la estructura del notebook.")
    else:
        logger.info("EXAMEN Estructura Correcta: El notebook sigue la estructura esperada.")
        
        
        
        
  ################# EXTRAER INFORMACIÓN EXAMEN ########################################################################################################      
        
        
        
def extrae_informacion_examen(examen_file, criterios_validos):
    """
    Extrae la información del examen desde un notebook Jupyter y verifica que los criterios estén en la lista de criterios válidos.

    Parameters:
        examen_file (str): Ruta del archivo del notebook de examen.
        criterios_validos (dict_keys): Las claves de un diccionario de criterios válidos.

    Returns:
        dict: Un diccionario con el 'contexto' del examen y una lista de 'ejercicios', 
              donde cada ejercicio tiene 'enunciado', 'criterios', y 'solucion'.

    Raises:
        RuntimeError: Si ocurre un error crítico durante la extracción de información,
                      o si un criterio no está en la lista de criterios válidos.
    """
    examen_info = {
        "contexto": "",
        "ejercicios": []
    }

    contexto_detectado = False
    ejercicio_num = 0
    se_espera_solucion = False
    solucion_detectada = False
    ejercicio_info = {}

    # Logger específico para esta función
    logger = logging.getLogger('extrae_informacion_examen')
    log_stream = logging.StreamHandler()
    logger.addHandler(log_stream)
    logger.setLevel(logging.DEBUG)

    try:
        # Leer el notebook
        with open(examen_file, 'r', encoding='utf-8') as f:
            notebook = nbformat.read(f, as_version=4)
            
    except FileNotFoundError:
        error_msg = f"El archivo {examen_file} no se encontró."
        logger.critical(error_msg)
        raise RuntimeError(error_msg)
    except Exception as e:
        error_msg = f"Error al leer el archivo {examen_file}: {e}"
        logger.critical(error_msg)
        raise RuntimeError(error_msg)

    try:
        # Procesar las celdas del notebook
        for cell in notebook.cells:
            if cell.cell_type == 'markdown':
                cell_content = cell['source'].strip()

                # Extraer Contexto
                if cell_content.startswith("## Contexto"):
                    examen_info["contexto"] = cell_content.split("## Contexto:")[-1].strip()
                    contexto_detectado = True
    
                # Extraer Enunciado y Criterios del Ejercicio
                if cell_content.startswith("## Ejercicio"):
                    if ejercicio_num > 0:
                        examen_info["ejercicios"].append(ejercicio_info)

                    ejercicio_num += 1
                    se_espera_solucion = True
                    solucion_detectada = False
                    ejercicio_info = {
                        "enunciado": cell_content.split("## Ejercicio")[1].split("Criterios:")[0].strip(),
                        "criterios": [],
                        "solucion": []
                    }
                    criterios_text = cell_content.split("Criterios:")[-1].strip()
                    criterios_list = [criterio.strip() for criterio in criterios_text.split(",")]
                    
                    # Verificar que los criterios estén en la lista de criterios válidos
                    for criterio in criterios_list:
                        criterio_sin_arrobas = criterio.replace("@@", "")
                        if criterio_sin_arrobas not in criterios_validos:
                            error_msg = f"Criterio no válido encontrado: {criterio} en el ejercicio {ejercicio_num}."
                            logger.critical(error_msg)
                            raise RuntimeError(error_msg)
                    
                    ejercicio_info["criterios"] = criterios_list

            elif cell.cell_type == 'code' and se_espera_solucion:
                # Extraer solución de las celdas de código
                ejercicio_info["solucion"].append(cell['source'].strip())
                solucion_detectada = True

        # Agregar la información del último ejercicio si no ha sido añadido
        if ejercicio_num > 0 and ejercicio_info:
            examen_info["ejercicios"].append(ejercicio_info)

        if not contexto_detectado:
            error_msg = "No se detectó un '## Contexto' en el notebook."
            logger.critical(error_msg)
            raise RuntimeError(error_msg)

        logger.info("La información del examen ha sido extraída correctamente.")

    except Exception as e:
        error_msg = f"Error inesperado al procesar el archivo {examen_file}: {e}"
        logger.critical(error_msg)
        raise RuntimeError(error_msg)

    return examen_info




################# PROCESA RESPUESTA ALUMNO ################################################################################################################




def procesa_respuestas_alumno(alumno_file, numero_ejercicios):
    """
    Procesa un notebook de respuestas de un alumno para extraer las soluciones a los ejercicios.

    Parameters:
        alumno_file (str): Ruta del archivo del notebook de respuesta del alumno.
        numero_ejercicios (int): Número esperado de ejercicios en el examen.

    Returns:
        dict: Un diccionario con el 'nombre_alumno' y una lista de 'ejercicios', donde cada ejercicio
              tiene su 'enunciado' y 'solucion' (o un valor indicando que no fue respondido).
    """
    # Extraer el nombre del alumno desde el nombre del archivo
    nombre_alumno = os.path.splitext(os.path.basename(alumno_file))[0]

    # Crear un logger específico para la función
    logger = logging.getLogger(f'procesa_respuestas_alumno_{nombre_alumno}')
    log_stream = logging.StreamHandler()
    logger.addHandler(log_stream)
    logger.setLevel(logging.DEBUG)

    # Iniciar el log indicando el inicio del procesamiento
    logger.info(f"Procesando respuestas del alumno {nombre_alumno}")

    respuestas_alumno = {
        "nombre_alumno": nombre_alumno,
        "ejercicios": []
    }

    ejercicio_num = 0
    se_espera_solucion = False
    ejercicio_info = {}

    try:
        # Leer el notebook
        with open(alumno_file, 'r', encoding='utf-8') as f:
            notebook = nbformat.read(f, as_version=4)

    except FileNotFoundError:
        error_msg = f"Archivo no encontrado para el alumno {nombre_alumno}: {alumno_file}"
        logger.error(error_msg)
        return {"error": error_msg}
    except Exception as e:
        error_msg = f"Error al leer el notebook para el alumno {nombre_alumno}: {e}"
        logger.error(error_msg)
        return {"error": f"Error al leer el notebook: {str(e)}"}

    try:
        # Procesar las celdas del notebook
        for cell in notebook.cells:
            if cell.cell_type == 'markdown':
                cell_content = cell['source'].strip()

                # Identificar un nuevo ejercicio basado en el enunciado
                if cell_content.startswith("## Ejercicio"):
                    # Si estábamos esperando una solución y no se añadió código, marcamos como no respondido
                    if se_espera_solucion:
                        if not ejercicio_info.get("solucion") or all(
                            not line.strip() or line.strip().startswith("#") 
                            for solution in ejercicio_info["solucion"] 
                            for line in solution.splitlines()):
                            ejercicio_info["solucion"] = "No respondido"
                        respuestas_alumno["ejercicios"].append(ejercicio_info)

                    ejercicio_num += 1
                    se_espera_solucion = True
                    ejercicio_info = {
                        "enunciado": cell_content.split("## Ejercicio")[1].strip(),
                        "solucion": []
                    }

            elif cell.cell_type == 'code' and se_espera_solucion:
                # Procesar celdas de código para una solución
                codigo = cell['source'].strip()
                if codigo:
                    ejercicio_info["solucion"].append(codigo)

        # Agregar la información del último ejercicio si no ha sido añadido
        if se_espera_solucion:
            if not ejercicio_info.get("solucion") or all(
                not line.strip() or line.strip().startswith("#") 
                for solution in ejercicio_info["solucion"] 
                for line in solution.splitlines()):
                ejercicio_info["solucion"] = "No respondido"
            respuestas_alumno["ejercicios"].append(ejercicio_info)

        # Verificar que el número de ejercicios y soluciones coincida con el número esperado
        num_ejercicios = len(respuestas_alumno["ejercicios"])
        num_soluciones = sum(1 for ejercicio in respuestas_alumno["ejercicios"] if ejercicio["solucion"] != "No respondido")

        if num_ejercicios != numero_ejercicios:
            error_msg = f"El número de ejercicios en el notebook ({num_ejercicios}) no coincide con el número esperado ({numero_ejercicios}) para el alumno {nombre_alumno}."
            logger.error(error_msg)
            return {"error": error_msg}

        if num_soluciones != numero_ejercicios:
            error_msg = f"El número de soluciones en el notebook ({num_soluciones}) no coincide con el número esperado de ejercicios ({numero_ejercicios}) para el alumno {nombre_alumno}."
            logger.error(error_msg)
            return {"error": error_msg}

        logger.info(f"Respuestas del alumno {nombre_alumno} procesadas correctamente.")

    except Exception as e:
        error_msg = f"Error inesperado al procesar el archivo {alumno_file} para el alumno {nombre_alumno}: {e}"
        logger.error(error_msg)
        return {"error": f"Error inesperado: {str(e)}"}

    return respuestas_alumno



####################### COMPRUEBA EJECUCIÓN ####################################################################################################################



def comprueba_ejecucion(respuestas_alumno):
    """
    Comprueba si las soluciones de los ejercicios se ejecutan sin errores y añade el estado
    y el mensaje de error (si corresponde) al diccionario original de respuestas del alumno.

    Parameters:
        respuestas_alumno (dict): Diccionario que contiene las respuestas del alumno y sus soluciones.
        
    Returns:
        dict: El mismo diccionario 'respuestas_alumno' con los campos 'estado' y 'mensaje_de_error'
              añadidos a cada ejercicio.

    Raises:
        RuntimeError: Si ocurre un error crítico durante la ejecución de la función.
    """
    # Configurar un logger específico para el alumno
    nombre_alumno = respuestas_alumno["nombre_alumno"]
    logger_detallado = logging.getLogger(f'detallado_{nombre_alumno}')

    try:
        for i, ejercicio in enumerate(respuestas_alumno["ejercicios"], 1):
            solucion = ejercicio["solucion"]

            if solucion == "No respondido":
                ejercicio["estado"] = "No respondido"
                ejercicio["mensaje_de_error"] = None
            else:
                # Redirigir la salida estándar a un objeto StringIO para capturarla
                original_stdout = sys.stdout
                sys.stdout = io.StringIO()

                # Reemplazar plt.show con una función que no hace nada
                original_show = plt.show
                plt.show = lambda *args, **kwargs: None

                try:
                    # Crear un diccionario para almacenar el entorno de ejecución
                    entorno_ejecucion = {}
                    # Concatenar todos los bloques de código en una sola cadena
                    codigo_completo = "\n".join(solucion)
                    # Ejecutar el código completo en el entorno de ejecución
                    exec(codigo_completo, entorno_ejecucion)
                    ejercicio["estado"] = "Correcto"
                    ejercicio["mensaje_de_error"] = None

                except Exception as e:
                    # Si hay un error en la ejecución, marcar el estado como "Error" y guardar el mensaje de error
                    ejercicio["estado"] = "Error"
                    ejercicio["mensaje_de_error"] = str(e)
                    logger_detallado.error(f"Error en la ejecución del Ejercicio {i} para {nombre_alumno}: {e}")
                finally:
                    # Cerrar todas las figuras abiertas para evitar la acumulación de memoria
                    plt.close('all')
                    # Restaurar plt.show a su comportamiento original
                    plt.show = original_show
                    # Restaurar la salida estándar original
                    sys.stdout = original_stdout

    except Exception as e:
        error_msg = f"Error inesperado en comprueba_ejecucion para el alumno {nombre_alumno}: {e}"
        logging.error(error_msg)
        raise RuntimeError(error_msg)

    return respuestas_alumno


################# EVALUAR CON CHATGPT ####################################################################################################################


def evaluar_con_chatgpt(contexto, codigo, enunciado, criterios_texto, prompt_template):
    """
    Evalúa el código de un ejercicio utilizando el modelo GPT-4 de OpenAI.

    Parameters:
        contexto (str): El contexto del examen.
        codigo (str): El código del ejercicio a evaluar.
        enunciado (str): El enunciado del ejercicio.
        criterios_texto (str): Texto que contiene los criterios específicos para este ejercicio.
        prompt_template (str): El template del prompt con marcadores para la descripción y el código.

    Returns:
        dict: Un diccionario con la puntuación y comentario del ejercicio.
    """
    # Crear un logger específico para la función
    logger = logging.getLogger(f'evaluar_con_chatgpt')
    log_stream = logging.StreamHandler()
    logger.addHandler(log_stream)
    logger.setLevel(logging.DEBUG)

    # Insertar los valores en el prompt, incluyendo los criterios específicos para este ejercicio
    prompt = prompt_template.format(contexto=contexto, enunciado=enunciado, codigo=codigo, criterios=criterios_texto)

    try:
        cliente = OpenAI()
        response = cliente.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a programming teaching assistant evaluating student code."},
                {"role": "user", "content": prompt}
            ]
        )

    except RateLimitError as e:
        error_msg = f"Rate limit error: {e}"
        logger.error(error_msg)
        return {"error": error_msg}
    except OpenAIError as e:
        error_msg = f"OpenAI API error: {e}"
        logger.error(error_msg)
        return {"error": error_msg}
    except Exception as e:
        error_msg = f"General error: {e}"
        logger.error(error_msg)
        return {"error": error_msg}

    # Extraer la respuesta de ChatGPT
    try:
        evaluacion = response.choices[0].message.content
    except (KeyError, IndexError) as e:
        error_msg = f"Error al procesar la respuesta de la API: {e}"
        logger.error(error_msg)
        return {"error": error_msg}

    return evaluacion



################# EVALUAR EJERCICIOS ####################################################################################################################



def evaluar_ejercicios(info_examen, resultados_alumno, prompt_file, criterios_file):
    """
    Evalúa los ejercicios utilizando GPT-4 y devuelve las notas y comentarios para cada ejercicio.
    
    Parameters:
        info_examen (dict): Diccionario con el contexto del examen y los enunciados de los ejercicios, obtenido de `extrae_informacion_examen`.
        resultados_alumno (dict): Diccionario con las soluciones y estado de ejecución de los ejercicios, obtenido de `comprueba_ejecucion`.
        prompt_file (str): Ruta del archivo de texto que contiene el prompt.
        criterios_file (str): Ruta del archivo de texto que contiene los criterios.

    Returns:
        dict: Un diccionario con las evaluaciones de cada ejercicio.

    Raises:
        RuntimeError: Si ocurre un error crítico durante la evaluación.
    """
    # Crear un logger específico para la función
    logger = logging.getLogger(f'evaluar_ejercicios')
    log_stream = logging.StreamHandler()
    logger.addHandler(log_stream)
    logger.setLevel(logging.DEBUG)

    evaluaciones = {}

    try:
        # Leer el prompt desde el archivo
        with open(prompt_file, 'r', encoding='utf-8') as file:
            prompt_template = file.read()

        # Cargar los criterios desde el archivo
        criterios_info = cargar_criterios(criterios_file)

        # Iterar sobre los ejercicios y sus resultados
        for i, (ejercicio_info, resultado_alumno) in enumerate(zip(info_examen['ejercicios'], resultados_alumno['ejercicios']), start=1):
            enunciado = ejercicio_info['enunciado']
            codigo = resultado_alumno['solucion']
            criterios_nombres = ejercicio_info.get('criterios', [])

            try:
                # Generar el texto de los criterios para el prompt
                criterios_texto = "\n\n".join([
                    f"**{criterio.strip('@')}**\nDescripción: {criterios_info[criterio.strip('@')]['descripcion']}\nEjemplo: {criterios_info[criterio.strip('@')]['ejemplo']}"
                    for criterio in criterios_nombres if criterio.strip('@') in criterios_info
                ])
            except KeyError as e:
                logger.error(f"Error en el formato de los criterios: {e}")
                raise RuntimeError(f"Error en el formato de los criterios: {e}")

            # Redirigir la salida estándar a un objeto StringIO para capturarla
            original_stdout = sys.stdout
            sys.stdout = io.StringIO()

            try:
                if resultado_alumno['estado'] == "No respondido":
                    logger.warning(f"El ejercicio {i} no fue respondido por el alumno.")
                    evaluaciones[f'Ejercicio {i}'] = (
                        "**Puntuaciones**: [0]\n"
                        "**Comentarios**: [\"El ejercicio no fue respondido.\"]\n"
                        "**Comentario General**: [\"El alumno no proporcionó ninguna solución para este ejercicio.\"]"
                    )
                elif resultado_alumno['estado'] == "Error":
                    logger.warning(f"El ejercicio {i} contiene un error en la ejecución.")
                    evaluaciones[f'Ejercicio {i}'] = (
                        "**Puntuaciones**: [0]\n"
                        f"**Comentarios**: [\"Error en la ejecución: {resultado_alumno['mensaje_de_error']}\"]\n"
                        "**Comentario General**: [\"El código presentado contiene errores que impiden su correcta ejecución.\"]"
                    )
                else:
                    resultado = evaluar_con_chatgpt(
                        info_examen['contexto'], codigo, enunciado, criterios_texto, prompt_template
                    )
                    evaluaciones[f'Ejercicio {i}'] = resultado

            except Exception as e:
                error_msg = f"Error al evaluar el ejercicio {i} para el alumno: {e}"
                logger.error(error_msg)
                evaluaciones[f'Ejercicio {i}'] = (
                    "**Puntuaciones**: [0]\n"
                    f"**Comentarios**: [\"Error durante la evaluación: {str(e)}\"]\n"
                    "**Comentario General**: [\"Ocurrió un error inesperado durante la evaluación del código.\"]"
                )

            finally:
                # Restaurar la salida estándar original
                sys.stdout = original_stdout

    except Exception as e:
        error_msg = f"Error inesperado en evaluar_ejercicios: {e}"
        logger.error(error_msg)
        raise RuntimeError(error_msg)

    finally:
        # Cerrar y remover el handler del logger
        log_stream.close()
        logger.removeHandler(log_stream)

    return evaluaciones



################# EXTRAE RESULTADOS ####################################################################################################################


def extraer_resultados(resultados, nombre_alumno):
    """
    Extrae las puntuaciones, comentarios y comentarios generales de una estructura de resultados para todos los ejercicios.

    Parámetros:
    resultados (dict): Diccionario con las evaluaciones para todos los ejercicios.
    nombre_alumno (str): Nombre del alumno.

    Retorna:
    dict: Un diccionario con los resultados estructurados para todos los ejercicios.
    """
    # Crear un logger específico para este alumno
    logger = logging.getLogger(f'extraer_resultados_{nombre_alumno}')
    log_stream = logging.StreamHandler()
    logger.addHandler(log_stream)
    logger.setLevel(logging.ERROR)
    
    # Inicializar el diccionario para almacenar los resultados
    resultados_detallados = {}

    try:
        # Recorrer cada ejercicio y extraer los resultados
        for ejercicio, evaluacion in resultados.items():
            try:
                comentarios = evaluacion  # El texto completo del resultado

                # Buscar el patrón para las puntuaciones usando una expresión regular
                puntuaciones_pattern = re.search(r'\*\*Puntuaciones\*\*:\s*(\[[^\]]*\])', comentarios)
                comentarios_pattern = re.search(r'\*\*Comentarios\*\*:\s*(\[.*?\])', comentarios, re.DOTALL)
                comentario_general_pattern = re.search(r'\*\*Comentario General\*\*:\s*(\[.*?\])', comentarios, re.DOTALL)

                if puntuaciones_pattern:
                    puntuaciones = eval(puntuaciones_pattern.group(1))
                else:
                    puntuaciones = []
                    logger.warning(f"No se encontraron puntuaciones para el ejercicio {ejercicio} de {nombre_alumno}")

                if comentarios_pattern:
                    comentarios_list = eval(comentarios_pattern.group(1))
                else:
                    comentarios_list = []
                    logger.warning(f"No se encontraron comentarios para el ejercicio {ejercicio} de {nombre_alumno}")

                if comentario_general_pattern:
                    comentario_general = eval(comentario_general_pattern.group(1))[0]  # Extrae el primer elemento
                else:
                    comentario_general = "No disponible"
                    logger.warning(f"No se encontró comentario general para el ejercicio {ejercicio} de {nombre_alumno}")

                resultados_detallados[ejercicio] = {
                    'puntuaciones': puntuaciones,
                    'comentarios': comentarios_list,
                    'comentario_general': comentario_general
                }

            except Exception as e:
                error_msg = f"Error procesando los resultados para '{nombre_alumno}' en el ejercicio '{ejercicio}': {e}"
                logger.error(error_msg)

    finally:
        # Cerrar y remover el handler del logger
        log_stream.close()
        logger.removeHandler(log_stream)

    return resultados_detallados



################# GENERAR INFORME PROFESOR ####################################################################################################################




def generar_informe_profesor(examen_info, resultados, nombre_examen, output_dir):
    """
    Genera un informe en PDF para el profesor que resume las calificaciones, comentarios y errores 
    para cada alumno, junto con gráficos de notas.

    Parameters:
        examen_info (dict): Diccionario que contiene el contexto del examen y los enunciados de los ejercicios.
        resultados (dict): Diccionario con los resultados de la evaluación para todos los alumnos.
        nombre_examen (str): Nombre del examen.
        output_dir (str): Directorio donde se guardará el informe en PDF.
    """

    try:
        # Crear el objeto PDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Agregar la portada
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, f"Informe Resumido de Evaluación - {nombre_examen}", ln=True, align='C')
        pdf.ln(20)
        
        # Inicializar las variables para las notas finales
        notas_finales = []
        notas_por_ejercicio = {f'Ejercicio {i+1}': [] for i in range(len(examen_info['ejercicios']))}

        # Resumen para cada alumno
        for alumno, resultado_alumno in resultados.items():
            if alumno == 'criterios':
                continue  # Saltar el diccionario de criterios, ya que no es un alumno
            
            pdf.add_page()
            pdf.set_font("Arial", "B", 14)
            
            # Calcular y añadir la nota final para el alumno al principio
            notas_ejercicios = []
            for i, ejercicio in enumerate(examen_info['ejercicios'], 1):
                resultado_ejercicio = resultado_alumno.get(f"Ejercicio {i}", {})
                puntuaciones = resultado_ejercicio.get('puntuaciones', [])
                if puntuaciones:
                    nota_media = sum(puntuaciones) / len(puntuaciones)
                    notas_ejercicios.append(nota_media)
            
            if notas_ejercicios:
                nota_final = sum(notas_ejercicios) / len(notas_ejercicios)
                notas_finales.append(nota_final)
                pdf.cell(0, 10, f"Alumno: {alumno.split('.')[0]} - Nota Final: {nota_final:.2f}", ln=True)
            else:
                pdf.cell(0, 10, f"Alumno: {alumno.split('.')[0]}", ln=True)
            
            pdf.ln(10)  # Añadir espacio para evitar superposiciones
            
            # Resumen de cada ejercicio
            for i, ejercicio in enumerate(examen_info['ejercicios'], 1):
                resultado_ejercicio = resultado_alumno.get(f"Ejercicio {i}", {})
                puntuaciones = resultado_ejercicio.get('puntuaciones', [])
                comentarios = resultado_ejercicio.get('comentarios', [])
                
                if puntuaciones and comentarios:
                    # Mostrar la puntuación y el comentario para cada criterio
                    for j, (puntuacion, comentario) in enumerate(zip(puntuaciones, comentarios)):
                        pdf.set_font("Arial", "", 12)
                        criterio_nombre = examen_info['ejercicios'][i-1]['criterios'][j].strip('@')
                        pdf.cell(0, 10, f"{criterio_nombre}: Puntuación - {puntuacion}", ln=True)
                        pdf.set_font("Arial", "I", 12)
                        pdf.multi_cell(0, 10, f"Comentario - {comentario}")
                    
                    # Mostrar la nota promedio del ejercicio
                    nota_media = sum(puntuaciones) / len(puntuaciones)
                    notas_por_ejercicio[f'Ejercicio {i}'].append(nota_media)
                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(0, 10, f"Nota del Ejercicio {i}: {nota_media:.2f}", ln=True)
                else:
                    # Si no hay puntuaciones/comentarios, se considera un error
                    pdf.set_font("Arial", "", 12)
                    pdf.cell(0, 10, "Errores:", ln=True)
                    pdf.multi_cell(0, 10, f"{resultado_ejercicio.get('comentarios', 'No disponible')}")
                
                pdf.ln(5)
            
        # Gráfica del histograma de notas finales
        plt.figure(figsize=(6, 4))
        plt.hist(notas_finales, bins=10, edgecolor='black')
        plt.title('Histograma de Notas Finales')
        plt.xlabel('Nota Final')
        plt.ylabel('Número de Alumnos')
        plt.tight_layout()
        hist_path_finales = os.path.join(output_dir, 'histograma_notas_finales.png')
        plt.savefig(hist_path_finales)
        plt.close()

        # Añadir el histograma de notas finales al PDF
        pdf.add_page()
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, 'Histograma de Notas Finales', ln=True)
        pdf.image(hist_path_finales, x=10, y=30, w=190)
        
        # Histograma de las notas y gráficas de criterios por ejercicio
        graficos_a_borrar = [hist_path_finales]
        for i, ejercicio in enumerate(examen_info['ejercicios'], 1):
            notas = notas_por_ejercicio[f'Ejercicio {i}']
            if notas:
                # Histograma de las notas
                plt.figure(figsize=(6, 4))
                plt.hist(notas, bins=10, edgecolor='black')
                plt.title(f'Histograma de Notas - Ejercicio {i}')
                plt.xlabel('Nota')
                plt.ylabel('Número de Alumnos')
                plt.tight_layout()
                img_path = os.path.join(output_dir, f'histograma_ejercicio_{i}.png')
                plt.savefig(img_path)
                plt.close()

                graficos_a_borrar.append(img_path)

                # Añadir el histograma al PDF
                pdf.add_page()
                pdf.set_font("Arial", "B", 14)
                pdf.cell(0, 10, f"Histograma de Notas - Ejercicio {i}", ln=True)
                pdf.image(img_path, x=10, y=30, w=190)

                # Gráfico de barras para las puntuaciones por criterio
                criterios = [criterio.strip('@') for criterio in examen_info['ejercicios'][i-1]['criterios']]

                # Calcular las puntuaciones por criterio manejando el caso de longitud desigual
                puntuaciones_criterio = []
                for j in range(len(criterios)):
                    puntuaciones = [
                        resultados[alumno][f'Ejercicio {i}']['puntuaciones'][j]
                        for alumno in resultados
                        if alumno != 'criterios' and f'Ejercicio {i}' in resultados[alumno] and
                        len(resultados[alumno][f'Ejercicio {i}']['puntuaciones']) > j
                    ]
                    puntuaciones_criterio.append(np.mean(puntuaciones))

                plt.figure(figsize=(6, 4))
                plt.bar(criterios, puntuaciones_criterio, color='blue', edgecolor='black')
                plt.title(f'Puntuaciones por Criterio - Ejercicio {i}')
                plt.xlabel('Criterio')
                plt.ylabel('Puntuación Promedio')
                plt.xticks(rotation=45, ha='right')  # Girar los nombres de los criterios 45 grados
                plt.tight_layout()
                img_path_criterios = os.path.join(output_dir, f'puntuaciones_criterios_ejercicio_{i}.png')
                plt.savefig(img_path_criterios)
                plt.close()

                graficos_a_borrar.append(img_path_criterios)

                # Añadir gráfico de puntuaciones por criterio al PDF
                pdf.add_page()
                pdf.image(img_path_criterios, x=10, y=30, w=190)
        
        # Guardar el PDF en el directorio de salida
        output_path = os.path.join(output_dir, f"Informe_Profesor_{nombre_examen}.pdf")
        pdf.output(output_path)
        logging.info(f"Informe resumido generado en: {output_path}")

        # Eliminar los archivos de gráficas temporales
        for grafico in graficos_a_borrar:
            if os.path.exists(grafico):
                os.remove(grafico)
                logging.info(f"Archivo gráfico temporal eliminado: {grafico}")

    except Exception as e:
        logging.error(f"Error al generar el informe del profesor: {str(e)}")
        
        
     ################# GENERAR INFORME PDF ALUMNOS ####################################################################################################################   
        
        
        
def generar_informe_pdf_alumnos(examen_info, resultados, nombre_examen, output_dir):
    """
    Genera informes en PDF para todos los alumnos basados en la evaluación de sus ejercicios.

    Parameters:
        examen_info (dict): Diccionario que contiene el contexto del examen y los enunciados de los ejercicios.
        resultados (dict): Diccionario con los resultados de la evaluación para todos los alumnos.
        nombre_examen (str): Nombre del examen.
        output_dir (str): Directorio donde se guardarán los informes en PDF.
    """

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for alumno in resultados.keys():
        if alumno == 'criterios':
            continue

        nombre_alumno = alumno.split('.')[0]

        try:
            # Crear el objeto PDF
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)

            # Agregar la portada
            pdf.add_page()
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, f"Informe de Evaluación de {nombre_alumno}", ln=True, align='C')
            pdf.ln(10)
            pdf.set_font("Arial", "", 12)
            pdf.cell(0, 10, f"Curso: Introducción a Python para Finanzas", ln=True, align='C')
            pdf.cell(0, 10, f"Examen: {nombre_examen}", ln=True, align='C')
            pdf.cell(0, 10, f"Fecha: {pd.Timestamp('now').strftime('%d/%m/%Y')}", ln=True, align='C')
            pdf.ln(20)

            # Introducción
            pdf.set_font("Arial", "", 12)
            pdf.multi_cell(0, 10, examen_info['contexto'])
            pdf.ln(10)

            # Resumen de Evaluación Global
            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "Resumen de Evaluación Global", ln=True)
            pdf.set_font("Arial", "", 12)

            # Calcular la puntuación total y la nota media
            notas_ejercicios = []
            for i in range(1, len(examen_info['ejercicios']) + 1):
                resultado_ejercicio = resultados[alumno].get(f"Ejercicio {i}", {})
                puntuaciones = resultado_ejercicio.get('puntuaciones', [])
                if puntuaciones:
                    nota_media = sum(puntuaciones) / len(puntuaciones)
                    notas_ejercicios.append(nota_media)
            
            if notas_ejercicios:
                nota_final = sum(notas_ejercicios) / len(notas_ejercicios)
            else:
                nota_final = 0

            pdf.cell(0, 10, f"Puntuación Total: {nota_final:.2f}", ln=True)
            pdf.ln(5)

            # Comentarios generales según la nota final
            if nota_final >= 9:
                comentarios_generales = "Excelente desempeño. Sigue así para mantener este nivel."
            elif nota_final >= 7:
                comentarios_generales = "Buen desempeño, pero hay algunas áreas que podrían beneficiarse de más práctica y revisión."
            elif nota_final >= 5:
                comentarios_generales = "Desempeño aceptable, pero es necesario trabajar más en ciertos aspectos."
            elif nota_final > 0:
                comentarios_generales = "El desempeño es bajo. Se recomienda revisar los conceptos básicos y practicar más."
            else:
                comentarios_generales = "El alumno no ha completado satisfactoriamente el examen. Es necesario un repaso completo de los temas cubiertos."

            pdf.multi_cell(0, 10, comentarios_generales)
            pdf.ln(10)

            # Detalle de Evaluación por Ejercicio
            for i, ejercicio in enumerate(examen_info['ejercicios'], 1):
                pdf.add_page()  # Cada ejercicio comienza en una nueva página
                pdf.set_font("Arial", "B", 12)

                # Calcular la puntuación del ejercicio
                resultado_ejercicio = resultados[alumno].get(f"Ejercicio {i}", {})
                puntuaciones = resultado_ejercicio.get('puntuaciones', [])
                if puntuaciones:
                    nota_media = sum(puntuaciones) / len(puntuaciones)
                else:
                    nota_media = 0

                # Título del ejercicio con la puntuación
                pdf.cell(0, 10, f"Ejercicio {i}: {ejercicio['enunciado'].splitlines()[0]}", ln=False)
                pdf.set_x(-50)  # Mueve el cursor para la puntuación hacia la derecha
                pdf.cell(0, 10, f"Puntuación: {nota_media:.2f}", ln=True, align='R')
                pdf.set_font("Arial", "", 12)

                # Enunciado completo
                pdf.set_font("Arial", "B", 12)
                pdf.cell(0, 10, "1. Enunciado:", ln=True)
                pdf.set_font("Arial", "", 12)
                pdf.multi_cell(0, 10, ejercicio['enunciado'])
                pdf.ln(5)

                # Criterios evaluados
                pdf.set_font("Arial", "B", 12)
                pdf.cell(0, 10, "2. Criterios Evaluados:", ln=True)
                pdf.set_font("Arial", "", 12)
                for j, criterio in enumerate(ejercicio['criterios']):
                    criterio_nombre = criterio.strip('@')
                    descripcion_criterio = resultados['criterios'][criterio_nombre]['descripcion']
                    puntuacion_criterio = puntuaciones[j] if j < len(puntuaciones) else "N/A"
                    pdf.multi_cell(0, 10, f"- {criterio_nombre} ({puntuacion_criterio}): {descripcion_criterio}")
                pdf.ln(5)

                if resultado_ejercicio:
                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(0, 10, "3. Comentarios:", ln=True)
                    pdf.set_font("Arial", "", 12)
                    for comentario in resultado_ejercicio['comentarios']:
                        pdf.multi_cell(0, 10, f"- {comentario}")
                    pdf.ln(5)

                    pdf.set_font("Arial", "B", 12)
                    pdf.cell(0, 10, "4. Comentario General:", ln=True)
                    pdf.set_font("Arial", "", 12)
                    pdf.multi_cell(0, 10, resultado_ejercicio['comentario_general'])
                    pdf.ln(10)
                else:
                    pdf.cell(0, 10, "No se encontraron resultados para este ejercicio.", ln=True)
                    pdf.ln(10)

            # Conclusión y Recomendaciones ajustadas al nivel del examen
            pdf.set_font("Arial", "B", 14)
            pdf.cell(0, 10, "Conclusión y Recomendaciones", ln=True)
            pdf.set_font("Arial", "", 12)
            
            if nota_final >= 9:
                conclusion = "Has demostrado un excelente dominio del material. Mantén este nivel de esfuerzo y sigue perfeccionando tus habilidades."
            elif nota_final >= 7:
                conclusion = "Buen trabajo, pero hay algunas áreas que podrían beneficiarse de más práctica y revisión."
            elif nota_final >= 5:
                conclusion = "El desempeño es aceptable, pero se recomienda revisar los temas en los que tuviste más dificultades para mejorar en futuras evaluaciones."
            elif nota_final > 0:
                conclusion = "El rendimiento en este examen indica que es necesario un repaso más exhaustivo de los temas cubiertos. Considera buscar apoyo adicional para resolver dudas."
            else:
                conclusion = "Es crucial que dediques tiempo a revisar todos los conceptos clave del curso. Considera buscar ayuda para entender mejor los temas."

            pdf.multi_cell(0, 10, conclusion)
            pdf.ln(10)

            # Guardar el PDF en el directorio de salida
            output_path = os.path.join(output_dir, f"{nombre_alumno}_informe.pdf")
            pdf.output(output_path)
            logging.info(f"Informe generado para {nombre_alumno} en: {output_path}")

        except Exception as e:
            logging.error(f"Error al generar el informe para {nombre_alumno}: {str(e)}")
            
            
            
################# GENERAR EXCEL RESULTADOS ####################################################################################################################


def generar_excel_resultados(resultados, examen_info, filename_agrupado='resultados_evaluacion_agrupado.xlsx', filename_detallado='resultados_evaluacion_detallado.xlsx'):
    # Estructuras para almacenar los datos de ambos archivos Excel
    filas_agrupado = []
    filas_detallado = []

    # Iterar sobre los resultados de cada alumno
    for alumno, ejercicios in resultados.items():
        if alumno == 'criterios':  # Ignorar la clave de criterios
            continue
        
        fila_agrupada = {'Alumno': alumno}
        fila_detallada = {'Alumno': alumno}
        notas = []

        # Iterar sobre los ejercicios del alumno
        for i, (ejercicio, contenido) in enumerate(ejercicios.items(), start=1):
            if ejercicio == 'NOTA FINAL':  # Saltar la nota final, ya que se calculará de nuevo
                continue

            # Verificar si el ejercicio contiene un error de ejecución
            if contenido == {'puntuaciones': [0], 'comentarios': ["Error en la ejecución: name 'volatility' is not defined"], 'comentario_general': 'El código presentado contiene errores que impiden su correcta ejecución.'}:
                # Si hay un error de ejecución, asignar 0 a todos los criterios
                puntuaciones = [0] * len(examen_info['ejercicios'][i-1]['criterios'])
                nota_media = 0
            else:
                # Calcular la nota media del ejercicio
                puntuaciones = contenido['puntuaciones']
                nota_media = np.mean(puntuaciones)
            
            fila_agrupada[ejercicio] = nota_media
            notas.append(nota_media)
            
            # Añadir la nota final del ejercicio
            fila_detallada[f'{ejercicio} - NOTA FINAL'] = nota_media
            
            # Iterar sobre los criterios evaluados en el ejercicio
            criterios_evaluados = examen_info['ejercicios'][i-1]['criterios']  # i-1 porque los ejercicios están basados en índice
            for criterio, puntuacion in zip(criterios_evaluados, puntuaciones):
                criterio_nombre = criterio.strip('@@')  # Quitar los "@@" si existen
                fila_detallada[f'{ejercicio} - {criterio_nombre}'] = puntuacion

        # Calcular la nota final como la media de las notas de los ejercicios
        nota_final = np.mean(notas)
        fila_agrupada['NOTA FINAL'] = nota_final
        fila_detallada['NOTA FINAL'] = nota_final

        filas_agrupado.append(fila_agrupada)
        filas_detallado.append(fila_detallada)

    # Crear los DataFrames y guardarlos en Excel
    df_agrupado = pd.DataFrame(filas_agrupado)
    df_agrupado.to_excel(filename_agrupado, index=False)
    print(f"Archivo agrupado guardado en {filename_agrupado}")

    df_detallado = pd.DataFrame(filas_detallado)
    df_detallado.to_excel(filename_detallado, index=False)
    print(f"Archivo detallado guardado en {filename_detallado}")


################# GUARDAR PROBLEMAS ####################################################################################################################


def guardar_problemas(problemas, output_dir):
    """
    Guarda la lista de problemas en un archivo de texto en el directorio de reports.

    Parameters:
        problemas (list): Lista de problemas encontrados durante la evaluación.
        output_dir (str): Directorio donde se guardará el archivo de problemas.
    """
    problemas_file = os.path.join(output_dir, 'problemas_entregas.txt')
    
    with open(problemas_file, 'w') as f:
        for problema in problemas:
            f.write(f"{problema}\n")
    
    print(f"Archivo de problemas guardado en: {problemas_file}")



################# PROCESA Y EVALUA NOTEBOOK ####################################################################################################################


def procesa_y_evalua_notebook(fich, directorio_entregas, examen_info, criterios_file, prompt_file, dir_log):
    """
    Procesa y evalúa un notebook individual.

    Parameters:
        fich (str): Nombre del archivo del notebook del alumno.
        directorio_entregas (str): Ruta del directorio donde se encuentran los notebooks.
        examen_info (dict): Información del examen.
        criterios (dict): Criterios de evaluación.
        prompt_file (str): Ruta del archivo de prompt.
        dir_log (str): Ruta del directorio donde se guardarán los logs de error.
        
    Returns:
        tuple: Nombre del alumno y resultados de la evaluación, o None si ocurre un error.
    """
    # Crear logger temporal para este notebook
    logger_temporal, log_stream = obtener_logger_temporal()

    try:
        
        # Procesar respuestas del alumno
        alumno_file = os.path.join(directorio_entregas, fich)
        respuestas_alumno = procesa_respuestas_alumno(alumno_file, len(examen_info['ejercicios']))
        
        # Comprobar la ejecución de las respuestas
        respuestas_evaluadas = comprueba_ejecucion(respuestas_alumno)
        
        # Evaluar las respuestas utilizando ChatGPT
        evaluaciones = evaluar_ejercicios(examen_info, respuestas_evaluadas, prompt_file, criterios_file)
        
        # Extraer los resultados detallados de la evaluación
        resultados_detallados = extraer_resultados(evaluaciones, fich)
        
        return fich, resultados_detallados
    
    except Exception as e:
        # Registrar el error en el log temporal
        logger_temporal.error(f"Error procesando el notebook {fich}: {e}")
        # Guardar el log detallado en caso de error
        guardar_log_error(log_stream, fich, dir_log)
        return None
    
    finally:
        # Cerrar el log temporal
        log_stream.close()
        
        
################# OBTENER LOGGER TEMPORAL ####################################################################################################################


# Función para obtener un logger temporal

def obtener_logger_temporal():
    log_stream = StringIO()
    handler = logging.StreamHandler(log_stream)
    logger = logging.getLogger(f'logger_temporal_{id(log_stream)}')
    logger.addHandler(handler)
    logger.setLevel(logging.ERROR)
    return logger, log_stream


################# GUARDAR LOG DE ERROR ####################################################################################################################


# Función para guardar logs de error en un archivo separado
def guardar_log_error(log_stream, nombre_notebook, dir_log):
    error_log_file = os.path.join(dir_log, f'log_errores_{nombre_notebook}.log')
    with open(error_log_file, 'w') as f:
        f.write(log_stream.getvalue())
        

################## CONFIGURAR LOGS ####################################################################################################################


def configurar_logs(directorio_raiz):
    dir_log = os.path.join(directorio_raiz, "logs")

    if os.path.exists(dir_log):
        # Si el directorio existe, eliminar los archivos de log antiguos
        for log_file in ['log_critico.log', 'log_detallado.log']:
            log_path = os.path.join(dir_log, log_file)
            if os.path.exists(log_path):
                os.remove(log_path)
    else:
        os.makedirs(dir_log)

    # Configurar log crítico global (mostrado en consola y en archivo)
    log_critico_path = os.path.join(dir_log, 'log_critico.log')
    handler_critico_file = logging.FileHandler(log_critico_path)
    handler_critico_console = logging.StreamHandler()

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            handler_critico_file,  # Registrar en archivo
            handler_critico_console  # Mostrar en consola
        ]
    )

    # Configurar log detallado global (solo en archivo)
    log_detallado_path = os.path.join(dir_log, 'log_detallado.log')
    logger_detallado = logging.getLogger('detallado')
    handler_detallado = logging.FileHandler(log_detallado_path)
    formatter_detallado = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler_detallado.setFormatter(formatter_detallado)
    logger_detallado.addHandler(handler_detallado)
    logger_detallado.setLevel(logging.DEBUG)

    return dir_log


################# MAIN ####################################################################################################################


def main(directorio_raiz, nombre_fich_examen, nombre_examen):
    try:
        # Configurar los logs
        dir_log = configurar_logs(directorio_raiz)
        
        # Log de inicio del proceso
        logging.info("Inicio del proceso de evaluación.")
        
        # Configurar las rutas basadas en el directorio raíz
        directorio_entregas = os.path.join(directorio_raiz, "entregas")
        directorio_examen = os.path.join(directorio_raiz, "examenes")
        directorio_reports = os.path.join(directorio_raiz, "reports")
        criterios_file = os.path.join(directorio_raiz, 'criterios.txt')
        fichero_examen = os.path.join(directorio_examen, nombre_fich_examen)
        fichero_res = 'resultados_evaluacion.xlsx'
        prompt_file = os.path.join(directorio_raiz, 'prompt.txt')

        # Crear el directorio de reports si no existe
        if not os.path.exists(directorio_reports):
            os.makedirs(directorio_reports)

        # Cargar criterios
        if not os.path.exists(criterios_file):
            raise FileNotFoundError(f"El archivo de criterios {criterios_file} no existe.")
        criterios = cargar_criterios(criterios_file)

        # Preprocesar el examen
        if not os.path.exists(fichero_examen):
            raise FileNotFoundError(f"El archivo de examen {fichero_examen} no existe.")
        examen_info = extrae_informacion_examen(fichero_examen, criterios.keys())

        # Verificar estructura del examen
        verifica_estructura_examen(fichero_examen)

        # Listar los notebooks entregados por los alumnos
        if not os.path.exists(directorio_entregas):
            raise FileNotFoundError(f"El directorio de entregas {directorio_entregas} no existe.")
        alumnos, ficheros = listar_notebooks(directorio_entregas)
        if not ficheros:
            raise FileNotFoundError("No se encontraron notebooks de alumnos en el directorio de entregas.")
        
        # Inicializar el diccionario de resultados
        resultados = {}
        resultados['criterios'] = criterios
        problemas = []  # Lista para almacenar los problemas con cada notebook

        # # Procesar y evaluar cada notebook en paralelo
        # resultados_alumnos = Parallel(n_jobs=-1, verbose = 13)(
        #     delayed(procesa_y_evalua_notebook)(fich, directorio_entregas, examen_info, criterios, prompt_file, dir_log) 
        #     for fich in tqdm(ficheros, desc="Procesando notebooks")
        # )
        
        # Procesar y evaluar cada notebook de forma secuencial
        resultados_alumnos = []
        for fich in tqdm(ficheros, desc="Procesando notebooks"):
            resultado = procesa_y_evalua_notebook(fich, directorio_entregas, examen_info, criterios_file, prompt_file, dir_log)
            resultados_alumnos.append(resultado)

        # Filtrar resultados exitosos y agregar al diccionario de resultados
        procesados_exitosamente = []
        for result in resultados_alumnos:
            if result is not None:
                alumno, res_extraido = result
                resultados[alumno] = res_extraido
                procesados_exitosamente.append(alumno)
            else:
                problemas.append(f"Un notebook no se pudo procesar correctamente.")

        # Generar los informes en PDF para cada estudiante
        generar_informe_pdf_alumnos(examen_info, resultados, nombre_examen, directorio_reports)
        
        # Comparar la lista de notebooks entregados con los informes generados
        informes_generados = os.listdir(directorio_reports)
        for alumno in alumnos:
            informe_esperado = f"{alumno.split('.')[0]}_informe.pdf"
            if informe_esperado not in informes_generados:
                problemas.append(f"{alumno}: No se generó un informe.")

        # Generar el informe para el profesor
        generar_informe_profesor(examen_info, resultados, nombre_examen, directorio_reports)

         # Generar el archivo Excel con los resultados de la evaluación
      
        # Ruta del archivo Excel con los resultados agrupados
        filename_agrupado = os.path.join(directorio_reports, fichero_res)

        # Ruta del archivo Excel con los resultados detallados
        filename_detallado = os.path.join(directorio_reports, 'resultados_evaluacion_detallado.xlsx')

        # Llamar a la función con las rutas especificadas
        generar_excel_resultados(resultados, examen_info, filename_agrupado=filename_agrupado, filename_detallado=filename_detallado)


 

        # Guardar el archivo con la lista de problemas
        guardar_problemas(problemas, directorio_reports)

        # Log de finalización exitosa
        logging.info("Proceso completado con éxito.")
    
    except Exception as e:
        error_msg = f"Error inesperado en el proceso: {e}"
        logging.critical(error_msg)
        print(error_msg, file=sys.stderr)
        


################# EJECUCIÓN ####################################################################################################################

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ejecuta la evaluación de exámenes.")
    parser.add_argument('--directorio_raiz', required=True, help="Directorio raíz donde se encuentran los archivos.")
    parser.add_argument('--nombre_fich_examen', required=True, help="Nombre del archivo de examen.")
    parser.add_argument('--nombre_examen', required=True, help="Nombre del examen para los informes.")
    
    args = parser.parse_args()
    
    main(args.directorio_raiz, args.nombre_fich_examen, args.nombre_examen)